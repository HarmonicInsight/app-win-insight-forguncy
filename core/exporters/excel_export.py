#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelエクスポートモジュール

解析結果をExcel形式で出力する。
"""

import os
import re
from datetime import datetime

from core.models import AnalysisResult


# =============================================================================
# バージョン情報
# =============================================================================
APP_VERSION = "1.1.0"
SUPPORTED_FORGUNCY_VERSIONS = ["9.x"]
VERSION_INFO = f"v{APP_VERSION} (Forguncy {', '.join(SUPPORTED_FORGUNCY_VERSIONS)} 対応)"


# =============================================================================
# openpyxl可用性チェック
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# =============================================================================
# ER図Mermaid生成
# =============================================================================
def generate_er_mermaid(tables: list) -> str:
    """ER図のMermaid記法を生成"""
    def sanitize(s):
        return re.sub(r'[^a-zA-Z0-9_\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF]', '_', s)

    lines = ['erDiagram']
    for table in tables:
        lines.append(f'  {sanitize(table.name)} {{')
        for col in table.columns[:10]:
            col_type = re.sub(r'[^a-z]', '', col.type.lower()) or 'string'
            pk = 'PK' if col.name.lower() == 'id' else ''
            lines.append(f'    {col_type} {sanitize(col.name)} {pk}'.strip())
        if len(table.columns) > 10:
            lines.append('    string more_columns "..."')
        lines.append('  }')

    for table in tables:
        for rel in table.relations:
            from_t = sanitize(table.name)
            to_t = sanitize(rel.target_table)
            rel_type = '}o--||' if 'Many' in rel.relation_type else '||--||'
            lines.append(f'  {from_t} {rel_type} {to_t} : "{rel.source_column}"')

    return '\n'.join(lines)


# =============================================================================
# Excel出力
# =============================================================================
def generate_excel_document(analysis: AnalysisResult, output_dir: str) -> str:
    """Excel形式で出力"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxlがインストールされていません。pip install openpyxl を実行してください。")

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()

    # ヘッダースタイル
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # サマリーシート
    ws_summary = wb.active
    ws_summary.title = 'サマリー'
    summary_data = [
        ['項目', '値'],
        ['プロジェクト名', analysis.project_name],
        ['テーブル数', analysis.summary.table_count],
        ['ページ数', analysis.summary.page_count],
        ['ワークフロー数', analysis.summary.workflow_count],
        ['サーバーコマンド数', analysis.summary.server_command_count],
        ['総カラム数', analysis.summary.total_columns],
        ['リレーション数', analysis.summary.total_relations],
        ['生成日', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['生成ツール', f'Forguncy Insight {VERSION_INFO}'],
    ]
    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40

    # テーブル一覧シート
    ws_tables = wb.create_sheet('テーブル一覧')
    table_headers = ['No.', 'テーブル名', 'フォルダ', 'カラム数', 'リレーション数']
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws_tables.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, t in enumerate(analysis.tables, 2):
        values = [row_idx - 1, t.name, t.folder or '-', len(t.columns), len(t.relations)]
        for col_idx, value in enumerate(values, 1):
            cell = ws_tables.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # カラム定義シート
    ws_columns = wb.create_sheet('カラム定義')
    col_headers = ['テーブル名', 'カラム名', 'データ型', '必須', 'ユニーク', 'デフォルト値']
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws_columns.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    for t in analysis.tables:
        for c in t.columns:
            values = [t.name, c.name, c.type, '○' if c.required else '', '○' if c.unique else '', c.default_value or '']
            for col_idx, value in enumerate(values, 1):
                cell = ws_columns.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
            row_idx += 1

    # ページ一覧シート
    ws_pages = wb.create_sheet('ページ一覧')
    page_headers = ['No.', 'ページ名', '種別', 'ボタン数', '数式数']
    for col_idx, header in enumerate(page_headers, 1):
        cell = ws_pages.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, p in enumerate(analysis.pages, 2):
        values = [row_idx - 1, p.name, 'マスターページ' if p.page_type == 'masterPage' else 'ページ', len(p.buttons), len(p.formulas)]
        for col_idx, value in enumerate(values, 1):
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # サーバーコマンドシート
    ws_cmds = wb.create_sheet('サーバーコマンド')
    cmd_headers = ['No.', 'コマンド名', 'フォルダ', 'パラメータ数', '処理行数']
    for col_idx, header in enumerate(cmd_headers, 1):
        cell = ws_cmds.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, c in enumerate(analysis.server_commands, 2):
        values = [row_idx - 1, c.name, c.folder or '-', len(c.parameters), len(c.commands)]
        for col_idx, value in enumerate(values, 1):
            cell = ws_cmds.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # ER図シート (テキスト形式)
    ws_er = wb.create_sheet('ER図(Mermaid)')
    ws_er.cell(row=1, column=1, value='以下をMermaid Live Editorに貼り付けてください')
    er_code = generate_er_mermaid(analysis.tables)
    ws_er.cell(row=3, column=1, value=er_code)

    file_path = os.path.join(output_dir, f'{analysis.project_name}_仕様書.xlsx')
    wb.save(file_path)
    return file_path


# =============================================================================
# 差分比較Excel出力
# =============================================================================
def generate_diff_excel(diff, old_name: str, new_name: str, output_dir: str) -> str:
    """差分比較結果をExcel形式で出力"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxlがインストールされていません。pip install openpyxl を実行してください。")

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()

    # スタイル定義
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    added_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    removed_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    modified_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    added_font = Font(color='006100')
    removed_font = Font(color='9C0006')
    modified_font = Font(color='9C6500')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # =========================
    # サマリーシート
    # =========================
    ws_summary = wb.active
    ws_summary.title = 'サマリー'

    summary_data = [
        ['差分比較レポート', ''],
        ['', ''],
        ['比較元（旧）', old_name],
        ['比較先（新）', new_name],
        ['生成日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['変更サマリー', ''],
        ['', ''],
        ['カテゴリ', '追加', '削除', '変更'],
    ]
    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 7]:
                cell.font = Font(bold=True, size=14)
            if row_idx == 9:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

    # サマリーデータ
    summary_rows = [
        ['テーブル', len(diff.added_tables), len(diff.removed_tables), len(diff.modified_tables)],
        ['ページ', len(diff.added_pages), len(diff.removed_pages), len(getattr(diff, 'modified_pages', []))],
        ['サーバーコマンド', len(diff.added_server_commands), len(diff.removed_server_commands), len(diff.modified_server_commands)],
    ]
    for row_idx, row in enumerate(summary_rows, 10):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 2 and value > 0:
                cell.fill = added_fill
                cell.font = added_font
            elif col_idx == 3 and value > 0:
                cell.fill = removed_fill
                cell.font = removed_font
            elif col_idx == 4 and value > 0:
                cell.fill = modified_fill
                cell.font = modified_font

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 10

    # =========================
    # テーブル変更シート
    # =========================
    ws_tables = wb.create_sheet('テーブル変更')
    table_headers = ['変更種別', 'テーブル名', 'フォルダ', '詳細']
    for col_idx, header in enumerate(table_headers, 1):
        cell = ws_tables.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    # 追加テーブル
    for t in diff.added_tables:
        values = ['追加', t.name, t.folder or '-', f'カラム数: {len(t.columns)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_tables.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = added_fill
            cell.font = added_font
        row_idx += 1

    # 削除テーブル
    for t in diff.removed_tables:
        values = ['削除', t.name, t.folder or '-', f'カラム数: {len(t.columns)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_tables.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = removed_fill
            cell.font = removed_font
        row_idx += 1

    # 変更テーブル
    for m in diff.modified_tables:
        details = []
        if m.get('added_columns'):
            details.append(f"追加カラム: {', '.join(c.name for c in m['added_columns'])}")
        if m.get('removed_columns'):
            details.append(f"削除カラム: {', '.join(c.name for c in m['removed_columns'])}")
        if m.get('modified_columns'):
            for mc in m['modified_columns']:
                details.append(f"{mc['name']}: {', '.join(mc['changes'])}")

        values = ['変更', m['name'], m['old'].folder or '-', '; '.join(details) if details else '構造変更']
        for col_idx, value in enumerate(values, 1):
            cell = ws_tables.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = modified_fill
            cell.font = modified_font
        row_idx += 1

    ws_tables.column_dimensions['A'].width = 10
    ws_tables.column_dimensions['B'].width = 30
    ws_tables.column_dimensions['C'].width = 20
    ws_tables.column_dimensions['D'].width = 60

    # =========================
    # ページ変更シート
    # =========================
    ws_pages = wb.create_sheet('ページ変更')
    page_headers = ['変更種別', 'ページ名', 'フォルダ', '詳細']
    for col_idx, header in enumerate(page_headers, 1):
        cell = ws_pages.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    # 追加ページ
    for p in diff.added_pages:
        values = ['追加', p.name, p.folder or '-', f'ボタン: {len(p.buttons)}, 数式: {len(p.formulas)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = added_fill
            cell.font = added_font
        row_idx += 1

    # 削除ページ
    for p in diff.removed_pages:
        values = ['削除', p.name, p.folder or '-', f'ボタン: {len(p.buttons)}, 数式: {len(p.formulas)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = removed_fill
            cell.font = removed_font
        row_idx += 1

    # 変更ページ
    for m in getattr(diff, 'modified_pages', []):
        details = []
        if m.get('added_buttons'):
            details.append(f"追加ボタン: {len(m['added_buttons'])}個")
        if m.get('removed_buttons'):
            details.append(f"削除ボタン: {len(m['removed_buttons'])}個")
        if m.get('added_formulas'):
            details.append(f"追加数式: {len(m['added_formulas'])}個")
        if m.get('removed_formulas'):
            details.append(f"削除数式: {len(m['removed_formulas'])}個")

        values = ['変更', m['name'], m['old'].folder or '-', '; '.join(details) if details else '内容変更']
        for col_idx, value in enumerate(values, 1):
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = modified_fill
            cell.font = modified_font
        row_idx += 1

    ws_pages.column_dimensions['A'].width = 10
    ws_pages.column_dimensions['B'].width = 40
    ws_pages.column_dimensions['C'].width = 20
    ws_pages.column_dimensions['D'].width = 50

    # =========================
    # サーバーコマンド変更シート
    # =========================
    ws_cmds = wb.create_sheet('サーバーコマンド変更')
    cmd_headers = ['変更種別', 'コマンド名', 'フォルダ', '詳細']
    for col_idx, header in enumerate(cmd_headers, 1):
        cell = ws_cmds.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    # 追加コマンド
    for c in diff.added_server_commands:
        values = ['追加', c.name, c.folder or '-', f'パラメータ: {len(c.parameters)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_cmds.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = added_fill
            cell.font = added_font
        row_idx += 1

    # 削除コマンド
    for c in diff.removed_server_commands:
        values = ['削除', c.name, c.folder or '-', f'パラメータ: {len(c.parameters)}']
        for col_idx, value in enumerate(values, 1):
            cell = ws_cmds.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = removed_fill
            cell.font = removed_font
        row_idx += 1

    # 変更コマンド
    for m in diff.modified_server_commands:
        details = []
        if m.get('added_parameters'):
            details.append(f"追加パラメータ: {', '.join(p.name for p in m['added_parameters'])}")
        if m.get('removed_parameters'):
            details.append(f"削除パラメータ: {', '.join(p.name for p in m['removed_parameters'])}")
        if m.get('commands_changed'):
            details.append("処理内容変更")

        values = ['変更', m['name'], m['old'].folder or '-', '; '.join(details) if details else '内容変更']
        for col_idx, value in enumerate(values, 1):
            cell = ws_cmds.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = modified_fill
            cell.font = modified_font
        row_idx += 1

    ws_cmds.column_dimensions['A'].width = 10
    ws_cmds.column_dimensions['B'].width = 40
    ws_cmds.column_dimensions['C'].width = 20
    ws_cmds.column_dimensions['D'].width = 50

    # 保存
    file_path = os.path.join(output_dir, f'差分比較_{old_name}_vs_{new_name}.xlsx')
    wb.save(file_path)
    return file_path
